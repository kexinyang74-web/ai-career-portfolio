"""数据加载层:CSV / Excel → 统一的 dict 行列表

设计:清洗逻辑(cleaner.py)不认识"文件格式"——
它只认 {列名: 值} 的 dict 行。
所以 CSV 和 Excel 只要在这一层变成同样的 dict 行,
后面的清洗流水线一个字都不用改(Day 26 设计留下的红利)。
"""
import csv

import openpyxl


def 读csv(路径):
    """读 CSV → [{列名: 值}, ...](DictReader,自动跳过空行)"""
    with open(路径, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def 读xlsx(路径):
    """读 Excel → [{列名: 值}, ...]

    openpyxl 读的三句话:
      1. load_workbook(路径)  打开文件
      2. wb.active            拿默认工作表(sheet)
      3. ws.iter_rows(values=True) 逐行取值,第一行是表头
    """
    wb = openpyxl.load_workbook(路径)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:                       # 第一行 = 表头(列名)
            表头 = list(row)
            continue
        if all(cell is None for cell in row):
            continue                     # 全空的 Excel 行直接跳过
        # Excel 空格子是 None,统一转成 ""(让下游和 CSV 行为一致)
        rows.append({表头[j]: ("" if cell is None else str(cell))
                     for j, cell in enumerate(row)})
    return rows
